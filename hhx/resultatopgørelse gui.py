from tkinter import *
from tkinter import ttk
from random import randint
import sys, os, socket, re

#from AutomatiskSearch import *
from Funktionsopdelt_resultatopgørelse import Funktionsopdelt_resultatopgørelse as FRO
from Artsopdelt_resulatopgørelse import Artsopdelt_resultatopgørelse as ARO
from Balance_handel import Balance as Balance_H

from settings import *

###########################Currentpath = sys.argv[0].rsplit('/', 1)[0] + '/'
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import openpyxl

#Colkcdskx-zm


class LicenseCode(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.master = master

        self.encrypt()
        self.master.title('┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳ ┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳ ')

        #################################### init window ##################################################
        canvas = Canvas(root)
        canvas.create_rectangle(0, 0, 500, 200, fill='lightblue')
        canvas.place(x=0, y=0, width=500, height=200)

        empty = Label(root, text='Indtast License Kode', relief="ridge", bg='white', height=2, width=20)
        empty.place(x=340, y=90)

        self.Input = Entry(root, text='Enter license code here', width=50)
        self.Input.place(x=20, y=100)

        label = Label(root, text='Kopier kode og send den til admin for at få din license kode', bg='lightgreen', relief="groove")
        label.place(x=10, y=10)

        label2 = Label(root, text='Code: ', bg='lightblue')
        label2.place(x=10, y=35)

        showCode = Entry(root, relief='groove')
        showCode.insert(END, encryptedMSG)
        showCode.place(x=55, y=35)

        # Buttons #
        ok = Button(root, text='check license', command=self.checkLicense)
        ok.place(x=20, y=130)

        quit = Button(root, text='Quit', command=self.master.destroy)
        quit.place(x=100, y=130)

    def encrypt(self):
        """Encrypts the hostname"""

        s = list(socket.gethostname())
        k = 5  # to move letter

        for i in range(len(s)):

            if s[i] == ' ':
                s[i] = '§' + str(randint(0, 9))
                continue

            elif not s[i].isalpha():
                continue

            crypt = ord(s[i]) + k
            if s[i].islower() and crypt > 122:
                crypt = 97 + (crypt - 97) % 26
            elif s[i].isupper() and crypt > 90:
                crypt = 65 + (crypt - 65) % 26
            s[i] = chr(crypt)

        global encryptedMSG
        encryptedMSG = ''.join(s)

        print('encryptedMSG:', encryptedMSG, 5)

    def decrypt(self, code):
        """Decrypts the secret (can't decrypt encryption only)"""

        s = list(code)
        k = 16

        for i in range(len(s)):

            if not s[i].isalpha():
                continue

            crypt = ord(s[i]) + k
            if s[i].islower() and crypt > 122:
                crypt = 97 + (crypt - 97) % 26
            elif s[i].isupper() and crypt > 90:
                crypt = 65 + (crypt - 65) % 26
            s[i] = chr(crypt)

        global hostname
        hostname = ''
        temp = False
        for i in range(len(s)):
            if s[i] == '§':
                temp = True
                continue

            if temp:
                hostname += ' '
                temp = False
            else:
                hostname += str(s[i])

        return hostname

    def checkLicense(self):
        code = self.decrypt(self.Input.get().strip())

        print(code, '=', socket.gethostname())

        if code == socket.gethostname():
            global licenseStatus
            licenseStatus = True
            self.master.destroy()
        else:
            error = Label(root, text='Forkert License Kode', bg='pink', height=2, width=20)
            error.place(x=340, y=90)


#root = Tk()
#root.geometry('{width}x{height}+{x}+{y}'.format(width=500, height=200, x=600, y=300))
#app = LicenseCode(root)
#root.mainloop()

# https://virksomheda.systime.dk/index.php?id=789#c6722


class StartWindow(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master=None)
        self.master = master
        self.master.title('  ┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳  Virksomhedsøkomoni  ┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳ ')

        Label(master, text="First").grid(row=0)
        Label(master, text="Second").grid(row=1, column=1)

        self.e1 = Entry(master)
        self.e2 = Entry(master)

        self.e1.grid(row=0, column=1)
        self.e2.grid(row=1, column=1)


        #self.pack(fill=BOTH, expand=200)






class Mainwindow(Frame):
    def __init__(self):
        Frame.__init__(self)
        self.master = master
        self.master.title('  ┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳  Virksomhedsøkomoni  ┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳┬ ┭ ┮ ┯ ┰ ┱ ┲ ┳ ')
        self.pack(fill=BOTH, expand=200)
        self.init_window()

    def init_window(self):
        #########################################
        #######      Menu     ###################
        menu = Menu(self.master)
        self.master.config(menu=menu)
        file = Menu(menu)
        file.add_command(label='Exit', command=self.master.destroy)
        menu.add_cascade(label='File', menu=file)
        edit = Menu(menu)
        edit.add_command(label='Undo')
        edit.add_command(label='show txt', command=self.showTxt)
        menu.add_cascade(label='Edit', menu=edit)
        #######     FONTS     ##########
        #helv36 = tkFont.Font(family="Helvetica", size=36, weight="bold")
        ####### Photo
        """photo = PhotoImage(file='C:\\Users\\sebastian\\Desktop\\Logo.png')
        label = Label(root, image=photo)
        label.place(x=0,y=0)"""
        #######   DRAWINGS    ##########

        canvas = Canvas(root)
        canvas.create_rectangle(0,0, 602, 106, fill='lightgreen' )
        #canvas.create_rectangle(0, 500, 200, 120, activefill='lightgreen', fill='black')
        canvas.place(x=0, y=0, width=602, height=400)

        #######    LABELS    ###########
        empty = Label(root, relief="ridge", bg='white', height=4, width=53)
        empty.place(x=10, y=320)
        label = Label(root, text='▼ Vælg Skabelon ▼', relief="groove", bg="lightblue", height=1,width=25)
        label.place(x=0, y=0)

        w = 30
        label2 = Label(root, text='Indast navnet til saldobalance-arket', relief='groove', width=w)
        label2.place(x=5, y=115)
        label3 = Label(root,text="Indtast cell koordinat til første og sidste konto i saldobalancen",relief='ridge',width=51, bg='lightblue')
        label3.place(x=230, y=112)

        self.first = Label(root, text='Første row:', relief='groove', width=15)
        self.first.place(x=230, y=137)
        self.last = Label(root, text='Sidste row:', relief='groove', width=15)
        self.last.place(x=415, y=137)

        w = 59
        FilePathLabel = Label(root, text='Indtast stien til filen hvori saldobalancen kan lokaliseres',
                              bg="lightgrey", relief='groove', width=w)
        FilePathLabel.place(x=184, y=0)
        SaveLocation = Label(root, text='Indtast stien til hvor filen skal gemmes', relief='groove',
                             bg='lightgrey', width=w)
        SaveLocation.place(x=184, y=50)

        #######  HOVERTEXT  ##########test
        """
        self.firstHover = Label(root, bg='lightgrey', text='test')
        self.firstHover.place(x=275, y=150)

        self.firstHover.bind("<Enter>", command=self.Enter)
        self.firstHover.bind("<Leave>", command=self.Leave)


        def Enter(self, event):
            self.firstHover.configure(text='Indtast row nummeret på den første konto nr.')

        def Leave(self, leave):
            self.firstHover.configure(text='')"""

        ####### RADIOBUTTONS ###########
        """
        v = IntVar()
        radio1 = Radiobutton(root, variable=v, text='Automatisk search (ustabilt)', value=2)
        radio1.place(x=200, y=200)"""

#
        #######  CHECKBOXES  ###########
        self.autoSearch = IntVar()

        self.autoBut = Checkbutton(root, variable=self.autoSearch, text='Automatisk search', command=self.hideEntryRows)
        self.autoBut.place(x=110, y=290)


        #####      ENTRY      ##########

        self.FileToLoad = Entry(root)
        self.FileToLoad.insert(END, 'C:/users/sebastian/desktop/opg_12_11')
        self.FileToLoad.place(x=192, y=21, width=400)
        self.SaveLocation = Entry(root)
        self.SaveLocation.insert(END, 'C:/users/sebastian/desktop/færdig.xlsx')                        #(sys.argv[0]).rsplit('/', 1)[0] + '/Færdig-Opgave.xlsx')
        self.SaveLocation.place(x=192, y=71, width=400)
        self.saldobalance = Entry(root)
        self.saldobalance.insert(END, 'Ark1')
        self.saldobalance.place(x=5, y=135, width=215)

        w = 60
        self.Entry_startrow = Entry(root)
        self.Entry_startrow.insert(END, 'A5')
        self.Entry_startrow.place(x=347, y=137, width=w)

        self.Entry_slutrow = Entry(root)
        self.Entry_slutrow.insert(END, 'A33')
        self.Entry_slutrow.place(x=532, y=137, width=w)

        #######    LISTBOX    #########

        self.skabelon = Listbox(root, height=5, width=30)
        self.skabelon.insert(END, 'Funktionsopdelt resultatopgørelse')
        for item in ['Artsopdelt resultatopgørelse', 'Balance', 'test', 'test2', 'test3']:
            self.skabelon.insert(END, item)
        self.skabelon.place(x=1, y=22)

        #######   SPINBOX   ##########    CURRENTLY NOT ACTIVE - test replace entry box

        spinval = StringVar()
        s = Spinbox(root, from_=1.0, to=100.0, textvariable=spinval)
        s = Spinbox(root, values=('test,','tesdfasdf2', '3', 'four', 'five bitch'))
        #s.place(x=100, y=300)

        ################################
        #######    BUTTONS    ##########

        startButton = Button(root, text='Start', relief='solid', width=14, height=2, bg='lightgreen',
                font=9, activebackground='green', command=self.START, overrelief='sunken')
        startButton.place(x=400, y=320)
        testButton = Button(root, command=self.doSomething, text='Do Something', width=14, overrelief='sunken')
        testButton.place(x=0, y=245)
        quitButton = Button(root, text="Quit", width=14, activebackground='red', overrelief='sunken', command=self.master.destroy)
        quitButton.place(x=0, y=285)
        ################################

    def START(self):
        """Start button - Laver skabelonen ud fra variablerne"""

        skabelon = self.skabelon.get(ACTIVE)
        saldobalance = self.saldobalance.get().strip()
        FilePath = self.Add_Xlsx_Extension(self.FileToLoad.get().strip())
        SaveLocation = self.Add_Xlsx_Extension(self.SaveLocation.get().strip())

        # Hvis autoSearch ikke er slået til så
        if not self.autoSearch.get():
            self.startCell = self.Entry_startrow.get().strip()
            self.slutCell = self.Entry_slutrow.get().strip()
            print('_' * 20 + '\nSkabelon: {}\nFilepath: {}\nSaveL: {}\nsaldob: {}\nstartrow: {}\nslutrow: {}'.format(skabelon, FilePath, SaveLocation, saldobalance, self.startCell, self.slutCell))


        # Kør function alt efter valgte skabelon
        if skabelon == 'Funktionsopdelt resultatopgørelse':
            if self.autoSearch.get(): # hvis automatisk er aktiveret

                # autoFindSaldobalance updaterer 'self.startCell' og 'self.slutCell' eller returner error 'status'
                status = self.autoFindSaldobalance(FileToLoad=FilePath, saldobalance=saldobalance)
                if status:
                    self.displayError(status)
                    return

            status = FRO(StartSaldoRow=self.startCell, EndSaldoRow=self.slutCell,
                                 FileToLoad=FilePath, SaveLocation=SaveLocation, saldobalance=saldobalance)

        elif skabelon == 'Artsopdelt resultatopgørelse':
            if self.autoSearch.get():

                status = self.autoFindSaldobalance(FileToLoad=FilePath, saldobalance=saldobalance)
                if status:
                    self.displayError(status)
                    return

            status = ARO(StartSaldoRow=self.startCell, EndSaldoRow=self.slutCell,
                         FileToLoad=FilePath, SaveLocation=SaveLocation, saldobalance=saldobalance)

        elif skabelon == 'Balance':
            if self.autoSearch.get():

                status = self.autoFindSaldobalance(FileToLoad=FilePath, saldobalance=saldobalance)
                if status:
                    self.displayError(status)
                    return

            status = Balance_H(startRow=self.startCell, slutRow=self.slutCell,
                     FileToLoad=FilePath, SaveLocation=SaveLocation, saldobalance=saldobalance)



        # Hvis en error (status) opstod - display it
        if status:
            self.displayError(status)
        else:
            succes = Label(root, text='Succesfully created the document', relief="ridge", bg="lightgreen", height=4, width=53)
            succes.place(x=10, y=320)


    def displayError(self, status):
        """ Displays the status error in the GUI """
        error = Label(root, text='ERROR: ' + status, relief="ridge", bg="pink", height=4, width=53)
        error.place(x=10, y=320)
    def Add_Xlsx_Extension(self, string):
        """Checks whether the string ends with xlsx, if it doesn't it adds the extension"""
        if string.endswith('.xlsx'):
            return string
        else:
            return string + '.xlsx'
    def CheckForError(self, status):
        """Checks if status is an error, and displays it"""
        if status:
            error = Label(root, text='ERROR: ' + status, relief="ridge", bg="pink", height=4, width=53)
            error.place(x=10, y=320)
            return True
    def changeCursor(self):
        succes = Label(root, text='testing ...', relief="ridge", bg="purple", height=4,
                       width=53)
        succes.place(x=10, y=320)
        root.config(cursor="xterm")
    def doSomething(self):
        skabelon = self.skabelon.get(ACTIVE)
        saldobalance = self.saldobalance.get()
        FilePath = self.Add_Xlsx_Extension(self.FileToLoad.get())
        SaveLocation = self.Add_Xlsx_Extension(self.SaveLocation.get())

        # check om startrow/slutrow er en integer
        try:
            startrow = int(self.Entry_startrow.get().strip())
            slutrow = int(self.Entry_slutrow.get().strip())
        except (TypeError, ValueError) as e:
            print(e)
            self.displayError('Først eller sidste row er ikke et tal\n(punktum og komma er ikke tilladt)')
    def hideEntryRows(self):
        """Hides 'Entry_startrow' when auto-search checkbox is clicked"""
        if self.autoSearch.get():
            self.notNeeded = Label(text='-  -  -  -  -    behøves ikke når auto search er slået til    -  -  -  -  -', width=51, bg='pink')
            self.notNeeded.place(x=230, y=137)
        else:
            self.notNeeded.destroy()
    def combineFuncs(self, *funcs):
        def combined_func(*args, **kwargs):
            for f in funcs:
                f(*args, **kwargs)
        return combined_func
    def showImg(self):
        load = Image.open('C:\\Users\\sebastian\\Desktop\\Logo.png')
        render = ImageTk.PhotoImage(load)
        img = Label(self, image=render)
        img.image = render
        img.place(x=0,y=0)
    def showTxt(self):
        text = Label(self, text='Hey there good lookin!')
        text.pack()
    def __isInt(self, input):
        """returns True if input is an integer"""
        try:
            int(input)
            return True
        except:
            return False
    def autoFindSaldobalance(self, FileToLoad, saldobalance):

        # Try load workbook and get saldobalance
        try:
            wb = openpyxl.load_workbook(FileToLoad)
            saldo = wb.get_sheet_by_name(saldobalance)

        except (FileNotFoundError, PermissionError) as e:
            return 'Could not find or load the specified document'
        except (NameError, KeyError) as e:
            return 'Kunne ikke lokalisere saldobalancen, check navnet på arket,\nhvori saldobalancen er lokaliseret.'



        def CheckSaldobalance():
            """Checker om saldobalancen er korrekt lokaliseret - returns 'self.sameRow' True/False
            afhængig af om deb/kredit er i samme row"""
            print('Kontonummer er i cell: {}{}'.format(letter, row))

            # check om deb/kredit står til højre for 'kontonr'; sameRow = True hvis samme row som 'kontonr'
            self.sameRow = False  # bool - er 'kontonr' i samme row som debet/kredit?
            debet = False
            kredit = False

            for r in range(2): # check row, da deb/kredit ikke altid er på samme row som kontonr
                for i in range(1, 7):
                    try:
                        string = saldo[chr(65 + i) + str(row+r)].value.lower().strip()
                        if string == 'debet':
                            debet = True
                        if string == 'kredit':
                            kredit = True
                    except AttributeError: # hvis cell value er int - kan ikke lower() - ignorere
                        pass
                if r == 0 and debet and kredit: # hvis deb/kred ikke var i samme row som 'kontonr'
                    self.sameRow = True


            if debet and kredit: # hvis debet og kredit blev fundet
                return True


        def findStartRow():
            print('findstartrow ----')
            dec = ord(letter)

            for i in range(1, 10):
                if (self.__isInt(saldo[letter + str(row+i)].value) and not self.__isInt(saldo[chr(dec+1) + str(row+i)].value)):
                    self.startCell = letter + str(row+i) # hvis kontonr er int, og kontonavn ikke er int
                    break

        def findSlutRow():
            startrow = int(self.startCell[1:])

            i = 1
            while True:
                if isInt(saldo[letter + str(startrow + i)].value):
                    i += 1
                else:
                    self.slutCell = letter + str(startrow + i)
                    return


        for column in range(saldo.max_column):  # chr(65)=A  chr(66)=B
            letter = chr(65 + column)

            for row in range(1, saldo.max_row):
                cell = letter + str(row)

                if saldo[cell].value:  # Hvis cell ikke er tom, check hvis der står 'kontonummer'
                    try:
                        result = re.search(r'^konto[-_ ]?(nummer|nr.?)$', saldo[cell].value.strip(),
                                           re.IGNORECASE).group()
                    except:
                        continue

                    if result:
                        print("Found 'kontonummer'")

                        if CheckSaldobalance():
                            findStartRow()
                            findSlutRow()
                            print('startrow ved cell:', self.startCell)
                            print('slutrow ved cell:', self.slutCell)
                            return
        else:
            return "Kunne ikke lokalisere saldobalancen i '{}' arket\n(Check om ark navnet er korrekt, evt. lokaliser den manuel)".format(saldobalance)




if __name__ == '__main__':
    #root = Tk()

    #root.geometry('{width}x{height}+{x}+{y}'.format(width=600, height=400, x=1400, y=300))
    #app = StartWindow(root)
#
    #root.mainloop()
#
#
    #exit()
    root = Tk()

    root.geometry('{width}x{height}+{x}+{y}'.format(width=600, height=400, x=1400, y=300))
    app = Mainwindow(root)

    root.mainloop()