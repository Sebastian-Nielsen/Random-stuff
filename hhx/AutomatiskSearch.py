
import re
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string




def autoFindSaldobalance(FileToLoad, saldobalance):

    # Try load workbook
    try:
        global wb
        wb = openpyxl.load_workbook(FileToLoad)
    except (FileNotFoundError, PermissionError) as e:
        return 'Could not find or load the specified document'

    # Try to get saldobalance
    try:
        saldo = wb.get_sheet_by_name(saldobalance)
    except (NameError, KeyError) as e:
        return 'Kunne ikke lokalisere saldobalancen, check navnet på arket,\nhvori saldobalancen er lokaliseret.'



    def CheckSaldobalance(letter, row):
        """Checker om saldobalancen er korrekt lokaliseret"""
        print('Kontonummer er i cell: {}{}'.format(letter, row))


        """TEST 1 - check hvis debet og kredit står til højre for 'kontonummer'"""
        def test1(row):
            debet = kredit = False
            row = str(row)

            for i in range(1, 7):
                try:
                    string = saldo[chr(65+i) + row].value.lower().strip()

                    if string == 'debet':
                        debet = True
                    if string == 'kredit':
                        kredit = True
                except AttributeError:
                    pass

            if debet and kredit:
                print('passed all tests #1')
                return True

        """TEST 2 - check hvis de to celler under 'kontonummer' er tal"""
        def test2():

            if str(saldo[letter + str(row+1)].value).isdigit and \
               str(saldo[letter + str(row+2)].value).isdigit:
                print('passed all tests #2')
                return True
            else:
                return False

        # hvis alle cases er True - find balancens slutRow                     #  TEST 3 - check kontonavne
        if test1(row) and test2():
            global AutoStartRow
            AutoStartRow = letter + str(row+1)
            return True

    def findSlutRow():

        def isInt(input):
            """returns True if input is an integer"""
            try:
                int(input)
                return True
            except:
                return False

        i = 3
        while True:

            if not isInt(saldo[letter + str(row + i)].value):
                global AutoSlutRow
                AutoSlutRow = letter + str(row + i -1)
                return

            i += 1



    for column in range(saldo.max_column):      # chr(65)=A  chr(66)=B
        letter = chr(65 + column)

        for row in range(1, saldo.max_row):
            cell = letter + str(row)
            print(saldo[cell].value)


            if saldo[cell].value:  # Hvis cell ikke er tom, check hvis der står 'kontonummer'
                try:
                    result = re.search(r'^konto[-_ ]?(nummer|nr.?)$', saldo[cell].value.strip(), re.IGNORECASE).group()
                except:
                    continue

                if result:
                    print("Found 'kontonummer'")
                    if CheckSaldobalance(letter, row):
                        print('kontonummer ved cell:', AutoStartRow)
                        findSlutRow()
                        print('slutRow ved cell:', AutoSlutRow)
                        return














