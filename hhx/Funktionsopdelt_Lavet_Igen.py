from settings import *
from openpyxl.styles import Font
import re

import openpyxl

def Funktionsopdelt_Automatisk(FileToLoad, SaveLocation, saldobalance, StartSaldoRow, EndSaldoRow):

    def PrintKontis():
        """Prints the data collected from: Debet and Kredit"""
        line = '_' * 30

        print(line)
        for i in debet:
            print('Debet:', i)
        print(line)
        for i in kredit:
            print('Kredit:', i)
        print(line)

    def Lokaliser_saldobalance():
        pass

    def GetData():
        """Collects the data (debet, kredit) from the FilePath xlsx document's saldobalance"""

        # Try to get saldobalance
        try:
            global wb
            wb = openpyxl.load_workbook(FileToLoad)
        except (FileNotFoundError, PermissionError) as e:
            print(e)
            return 'Could not find or load the specified document'

        # Try to get saldobalance
        try:
            saldo = wb.get_sheet_by_name(saldobalance)
        except (NameError, KeyError) as e:
            print(e)
            return 'Kunne ikke lokalisere saldobalancen'


        # Get data fra saldobalancen
        global kredit
        kredit = []
        global debet
        debet = []


        # Lokaliser saldobalance placering
        for column in range(saldo.max_column):
            # chr(65)=A  chr(66)=B  osv.
            letter = chr(65+column)

            for row in range(saldo.max_row):

                result = re.search(r'konto[-_ ]nummer', saldo[letter + row], re.IGNORECASE)

                if result:
                    print()
                    print('found one')
                    print(result)
                    print(letter, row)
                    print()
                    input()






        for i in range(StartSaldoRow, EndSaldoRow):
            if saldo['C' + str(i)].value == None:
                kredit.append([saldo['A' + str(i)].value,
                               saldo['B' + str(i)].value,
                               saldo['D' + str(i)].value,
                               saldo['D' + str(i)].coordinate])
            else:
                debet.append([saldo['A' + str(i)].value,
                              saldo['B' + str(i)].value,
                              saldo['C' + str(i)].value,
                              "='{}'!{}".format(saldobalance, saldo['C' + str(i)].coordinate)])







def Funktionsopdelt_resultatopgørelse(FileToLoad, SaveLocation, saldobalance, StartSaldoRow, EndSaldoRow):

    def PrintKontis():
        """Prints the data collected from: Debet and Kredit"""
        line = '_' * 30

        print(line)
        for i in debet:
            print('Debet:', i)
        print(line)
        for i in kredit:
            print('Kredit:', i)
        print(line)

    def GetData():
        """Collects the data (debet, kredit) from the FilePath xlsx document's saldobalance"""

        # Try load workbook
        try:
            global wb
            wb = openpyxl.load_workbook(FileToLoad)
        except (FileNotFoundError, PermissionError) as e:
            print(e)
            return 'Could not find or load the specified document'

        # Try to get saldobalance
        try:
            saldo = wb.get_sheet_by_name(saldobalance)
        except (NameError, KeyError) as e:
            print(e)
            return 'Kunne ikke lokalisere saldobalancen'

        # get data fra "saldobalance"
        global kredit
        kredit = []
        global debet
        debet = []

        for i in range(StartSaldoRow, EndSaldoRow):
            if saldo['C' + str(i)].value == None:
                kredit.append([saldo['A' + str(i)].value,
                               saldo['B' + str(i)].value,
                               saldo['D' + str(i)].value,
                               saldo['D' + str(i)].coordinate])
            else:
                debet.append([saldo['A' + str(i)].value,
                              saldo['B' + str(i)].value,
                              saldo['C' + str(i)].value,
                              "='{}'!{}".format(saldobalance, saldo['C' + str(i)].coordinate)])



    def CreateFROskabalon(row=3, title='Funktionsopd. Resultatopgørelse'):
        """Creates only the FRO-skabalon"""
        headlines = [
                     'Nettoomsætning',
                     'Produktionsomkostninger',
                     'Bruttofortjeneste',
                     'Distributionsomkostninger',
                     'Administrationsomkostninger',
                     'Resultat før primær drift',
                     'Finansielle indtægter',
                     'Finansielle omkostninger',
                     'Resultat før skat',
                     'Skat',
                     'Årets resultat'
                     ]

        # Lav ny "sheet" og gå ind på den
        wb.create_sheet(title=title, index=0)
        global doc
        doc = wb.get_sheet_by_name(title)


        # Overskrift
        doc['A1'] = 'Funktionsopdelt Resultatsopgørelse'
        doc['A1'].font = bold22Font


        # Lav resultatopgørelse skabelon
        for i in range(0, 11):
            doc['A' + str(row + i)] = headlines[i]

            # Gør headline bold
            if headlines[i] in ['Bruttofortjeneste', 'Resultat før primær drift', 'Resultat før skat', 'Årets resultat']:
                doc['A' + str(row + i)].font = bold10

            # Sæt note nr for enden
            else:
                temp = 'C' + str(row + i)
                if i == 1:
                    doc[temp] = 1
                elif i == 3:
                    doc[temp] = 2
                elif i == 4:
                    doc[temp] = 3





        ##### LAV OG UDFYLD NOTER ######
        noter = ['Produktionsomkostninger',
                 'Distributionsomkostninger',
                 'Administartionsomkostninger']

        # Da vi startede fra row: 3
        # - og vi har været igennem 11 headlines er vi nu kommet til row 14
        # - Vi adder 1 til fordi vi ønsker et mellemrum
        # - Det resultere i at vores current row nu er _15_
        ## noteSum er en list bestående af row nummeret af hver af noternes sum

        currentRow = 11 + row + 1
        startRow = 11 + row + 2
        global noteSum
        noteSum = []



        note = 1
        while note != 4:

            doc['A' + str(currentRow)] = 'Note {} - {}'.format(note, noter[note-1])
            doc['A' + str(currentRow)].font = bold10

            # Append data til noterne
            if note == 1:
                for konto in debet:
                    if 2000 <= konto[0] < 3000:
                        currentRow += 1
                        doc.append([konto[1], konto[3]])

            elif note == 2:
                for konto in debet:
                    if 3000 <= konto[0] < 4000:
                        currentRow += 1
                        doc.append([konto[1], konto[3]])

            elif note == 3:
                for konto in debet:
                    if 4000 <= konto[0] < 6000:
                        currentRow += 1
                        doc.append([konto[1], konto[3]])

            # add sum af alle tal i noten
            currentRow += 1
            print(currentRow, startRow)
            doc['A' + str(currentRow)] = noter[note-1] + ' i alt'
            doc['B' + str(currentRow)] = '=SUM(B{}:B{})'.format(startRow, currentRow - 1)

            # add underline under sum af note
            doc['A' + str(currentRow)].font = underline10
            doc['B' + str(currentRow)].font = underline10

            # append row nummeret
            noteSum.append(currentRow)


            note += 1
            startRow = currentRow + 2
            currentRow += 1

    def AddDataToFRO(row=3):
        """Adds data til FRO-skabelonen"""

        # formel for "netto omsætning"
        formula = '='
        for konto in kredit:
            if konto[0] < 1200:
                formula += '+' + saldobalance + '!' + konto[3]

        print(formula)
        doc['B' + str(row)] = formula


        # Add til "produktionsomkostninger
        row += 1
        doc['B' + str(row)] = "=B{}".format(noteSum[0])
        # Bruttofortjeneste
        row += 1
        doc['B' + str(row)] = "=B{}-B{}".format(row - 2, row - 1)
        # Distributionsomkostninger
        row += 1
        doc['B' + str(row)] = "=B{}".format(noteSum[1])
        # Administrationsomkostninger
        row += 1
        doc['B' + str(row)] = "=B{}".format(noteSum[2])
        # Resultat før primær drift
        row += 1
        doc['B' + str(row)] = "=B{}-B{}-B{}".format(row - 3, row - 2, row - 1)

        # Finansielle indtægter
        row += 1
        formula = '='

        for konto in kredit:
            if 4999 < konto[0] < 6000:
                formula += '+' + saldobalance + '!' + konto[3]

        if formula == '=':
            doc['B' + str(row)] = 0
        else:
            doc['B' + str(row)] = formula

        # Finansielle omkostninger
        row += 1
        formula = '='

        for konto in debet:
            if 4999 < konto[0] < 6000:
                formula += '+' + saldobalance + '!' + konto[3]

        if formula == '=':
            doc['B' + str(row)] = 0
        else:
            doc['B' + str(row)] = formula

        # Resultat før skat
        row += 1
        doc['B' + str(row)] = "=B{}+B{}-B{}".format(row - 3, row - 2, row - 1)

        # Skat
        row += 1
        formula = '='

        for konto in debet:
            if 6999 < konto[0] < 9000:
                formula += '+' + saldobalance + '!' + konto[3]

        if formula == '=':
            doc['B' + str(row)] = 0
        else:
            doc['B' + str(row)] = formula

        # Årets resultat
        row += 1
        doc['B' + str(row)] = "=B{}-B{}".format(row - 2, row - 1)

        doc.column_dimensions['A'].width = 40

        wb.save('asf')
        # Save file SaveLocation
        try:
            wb.save(SaveLocation)
        except (FileNotFoundError, PermissionError) as e:
            print(e)
            print('asdf----------')
            return 'Invalid save location'



    # status is used to return the error to the tkinter application

    status = GetData()

    if status:
        return status

    CreateFROskabalon()
    status = AddDataToFRO()

    if status:
        return status

    print('-Finished-')


