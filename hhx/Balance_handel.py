from settings import *
import openpyxl
from openpyxl.styles import Font

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

import re

# Forkortelse af "Balance" = BA


def Balance(FileToLoad, SaveLocation, saldobalance, startRow, slutRow):
    print(startRow)
    dec = ord(startRow[0])  # find dec (ascii_table - decimal) til letter   fx. A = 65 dec
    debetCol  = chr(ord(startRow[0]) + 2)
    kreditCol = chr(ord(startRow[0]) + 3)


    def PrintKontis():
        """Prints the data collected from: Debet and Kredit"""
        line = '_' * 30

        print(line)
        for i in debet:
            print('Debet:', i)
        print(line)
        for i in kredit:
            print('Kredit:', i)
        print(line)

    def GetData():
        """Collects the data (debet, kredit) from the FilePath xlsx document's saldobalance"""

        # Try load workbook
        try:
            global wb
            wb = openpyxl.load_workbook(FileToLoad)
        except (FileNotFoundError, PermissionError) as e:
            print(e)
            return 'Could not find or load the specified document'

        # Try to get saldobalance
        try:
            saldo = wb.get_sheet_by_name(saldobalance)
        except (NameError, KeyError) as e:
            print(e)
            return "Der findes ikke noget '{}' ark.\n(Den er case sensitive - dobbelttjek store og små bogstaver)".format(
                saldobalance)

        # get data fra "saldobalance"
        global konti
        konti = []

        for i in range(int(startRow[1:]), int(slutRow[1:])):  # Bogstaver hvis 'letter' er A:

            if saldo[chr(dec) + str(i)].value >= 10000:  # hvis konto-nr er over 10k add til list

                if saldo[chr(dec + 3) + str(i)].value != None:  # if cell C er None - så er det kredit

                    konti.append([saldo[chr(dec) + str(i)].value,                         # A
                                   saldo[chr(dec + 1) + str(i)].value,                      # B
                                   saldo[chr(dec + 3) + str(i)].coordinate[0],              # D
                                   "'{}'!{}".format(saldobalance, chr(dec + 3) + str(i))])  # D
                else:
                    konti.append([saldo[chr(dec) + str(i)].value,                        # A
                                  saldo[chr(dec + 1) + str(i)].value,                      # B
                                  saldo[chr(dec + 2) + str(i)].coordinate[0],              # C
                                  "'{}'!{}".format(saldobalance, chr(dec + 2) + str(i))])  # C

    def createBAskabelon(row=3, title='Balance'):
        """Laver selve 'balance' skabelonen, putter IKKE tal ind i den"""

        wb.create_sheet(title=title, index=0)
        global doc
        doc = wb.get_sheet_by_name(title)

        doc['A1'] = 'Balance - Handelsv.'
        doc['A1'].font = bold22Font

        doc['A3'] = 'Note'

        doc['B3'] = 'AKTIVER'
        doc['B4'] = 'ANLÆGSAKTIVER'

        doc['B12']  = 'OMSÆTNINGSAKTIVER'
        doc['B20'] = 'AKTIVER I ALT'

        doc['D3'] = 'Note'

        doc['E3'] = 'PASSIVER'
        doc['E4'] = 'EGENKAPITAL'

        doc['E12']  = 'GÆLDSFORPLIGTELSER'
        doc['E20'] = 'PASSIVER I ALT'

        for i in ('A3', 'B3', 'B4', 'B12', 'B20',
                  'D3', 'E3', 'E4', 'E12', 'E20'):
            doc[i].font = bold10

    def LavNote(i, noteRow, noteNR):
        """Laver en note - increases noteRow with 2, and noteNR with 1"""
        col1 = 'H'
        col2 = chr(ord('H')+1)


        print('in LavNote')


        doc[col1 + str(noteRow)] = "Note {} - {}".format(noteNR, konti[i][1]) # Overskrift
        doc[col1 + str(noteRow)].font = bold10
        noteRow  += 1
        doc[col1 + str(noteRow)] = konti[i][1] # 1. konto
        doc[col2 + str(noteRow)] = '=' + konti[i][3]
        noteRow  += 1
        doc[col1 + str(noteRow)] = konti[i+1][1] # 2. konto - akk. afskr.
        doc[col2 + str(noteRow)] = '=' + konti[i+1][3]
        noteRow  += 1
        doc[col1 + str(noteRow)] = "{} i alt".format(konti[i][1]) # sum af note
        doc[col2 + str(noteRow)] = "=+I{}-i{}".format(noteRow-2, noteRow-1)
        doc[col2 + str(noteRow)].font = underline10

        noteRow += 2
        noteNR += 1
        return noteRow, noteNR



    def addDataToBalance():
        """Add'er tal fra saldobalancen til BA skabelonen"""
        def placeSumFormula(col1, col2, row, startRow, slutRow, overskrift):
            """Add'er sum af tal til balance, regner ud optimal måde at sætte tallene op på"""
            print('in FORMULASUM')
            print(row)

            if row == slutRow:
                raise 'der er ikke mere plads til "sum"'

            if row + 1 < slutRow:  # hvis muligt at plasser sum med et felt mellemrum fra seneste tal - gør det.
                row += 1

            doc[col1 + str(row)] = overskrift

            doc[col2 + str(row)] = "=+SUM({}:{})".format(col2 + str(startRow), col2 + str(row - 1))
            doc[col2 + str(row)].font = underline10

            SumCellCoords.append(col2 + str(row))

        SumCellCoords = []
        noteRow = 1
        noteNR = 1

        # ANLÆGSAKTIVER
        row = 5
        i = 0

        while i != len(konti):
            if 11000 <= konti[i][0] < 12000:

                if konti[i][2] == debetCol: # konto er debet

                    if konti[i+1][2] == kreditCol and 11000 <= konti[i+1][0] < 12000: # hvis næste er kredit - LavNote
                        doc['A' + str(row)] = noteNR
                        noteRow, noteNR = LavNote(i, noteRow, noteNR)

                        # skriv i balance
                        doc['B' + str(row)] = konti[i][1] # navn
                        doc['C' + str(row)] = "=+I{}".format(str(noteRow-2)) # sum a note

                        row += 1
                        i += 2
                        continue
                    else:
                        doc['B' + str(row)] = konti[i][1]
                        doc['C' + str(row)] = '=' + konti[i][3]
                        row += 1
                        i += 1
                        continue
            i += 1

        placeSumFormula('B', 'C', row, 5, 12, 'Anlægsaktiver i alt')
        row += 1


        # OMSÆTNINGSAKTIVER
        row = 13
        i = 0

        while i != len(konti):
            if 12000 <= konti[i][0] < 13000:

                if konti[i+1][2] == kreditCol and 11000 <= konti[i+1][0] < 12000: # Lav Note
                    doc['A' + str(row)] = noteNR
                    noteRow, noteNR = LavNote(i, noteRow, noteNR)

                    # skriv i balance
                    doc['B' + str(row)] = konti[i][1]  # kontonavn
                    doc['C' + str(row)] = "=+I{}".format(str(noteRow - 2))  # sum af note

                    row += 1
                    i += 2
                    continue
                else:
                    doc['B' + str(row)] = konti[i][1]
                    doc['C' + str(row)] = '=' + konti[i][3]
                    row += 1
                    i += 1
                    continue
            i += 1

        placeSumFormula('B', 'C', row, 13, 20, 'Omsætningsaktiver i alt')

        # EGENKAPITAL
        row = 5
        i = 0

        while i != len(konti):
            if 13000 <= konti[i][0] < 14000:

                print(konti[i][1])
                if re.search(r'kapital[-_ ]?kont[oi]', konti[i][1], re.IGNORECASE): # kapitalkonto

                    print(konti[i+1][1])
                    if re.search(r'privat[-_ ]?forbrug', konti[i+1][1], re.IGNORECASE): # privatforbrug
                        doc['D' + str(row-1)] = noteNR
                        print('debug:', i, noteRow, noteNR)
                        noteRow, noteNR = LavNote(i, noteRow, noteNR)

                        # skriv i balance
                        print(row)
                        doc['F' + str(row-1)] = "=+I{}".format(str(noteRow - 2)) # sum af note

                        i += 2
                        continue
                    else:
                        doc['F' + str(row)] = konti[i][1]
                        i += 1
                        continue
            elif 14000 < konti[i][0]:
                break

            i += 1

        placeSumFormula('E', 'F', row, 4, 12, 'Omsætningsaktiver i alt')


        # GÆLDSFORPLIGTELSER
        """
        Jeg går ud fra at alle gældsforpligtelser hører til kredit.
        Derfor tager jeg ikke højde for debet, der bliver heller ikke lavet noter.
        """
        row = 13

        while i != len(konti):
            print(konti[i])
            doc['E' + str(row)] = konti[i][1]
            doc['F' + str(row)] = '=' + konti[i][3]

            row += 1
            i += 1

        placeSumFormula('E', 'F', row, 13, 20, 'Gældsforpligtelser i alt')



        # Passiver i alt
        doc['C20'] = "={}+{}".format(*SumCellCoords)
        # Aktiver i alt
        doc['F20'] = "={}+{}".format(SumCellCoords[2], SumCellCoords[3])

        setColumnWidth()

        try:
            wb.save(SaveLocation)
        except (FileNotFoundError, PermissionError) as e:
            print(e)
            return 'Invalid save location\n(If you are overwriting a document, remember to close it)'


    def setColumnWidth():
        # Set column width
        doc.column_dimensions['A'].width = 5
        doc.column_dimensions['B'].width = 40
        doc.column_dimensions['C'].width = 7
        doc.column_dimensions['D'].width = 5
        doc.column_dimensions['E'].width = 40
        doc.column_dimensions['f'].width = 7
        doc.column_dimensions['H'].width = 40
        doc.column_dimensions['I'].width = 10


    # status is used to return the error to the tkinter application
    status = GetData()

    if status:
        return status

    #PrintKontis()
    createBAskabelon()
    status = addDataToBalance()



    if status:
        return status

    print('-- finished --')

