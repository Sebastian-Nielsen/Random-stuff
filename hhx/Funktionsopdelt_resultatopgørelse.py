from settings import *
import openpyxl
from openpyxl.styles import Font

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell


# Forkortelse af "funktionsopdelte resultatopgørelse" = FRO


def Funktionsopdelt_resultatopgørelse(FileToLoad, SaveLocation, saldobalance, StartSaldoRow, EndSaldoRow):
    """Funktionsopdelt resultatopgørlse"""
    # StartSaldoRow = eg. 'A7'
    # EndSaldoRow = eg. 'A25'
    dec = ord(StartSaldoRow[0])  # find dec (ascii_table - decimal) til bogstaven

    def PrintKontis():
        """Prints the data collected from: Debet and Kredit"""
        line = '_' * 30

        print(line)
        for i in debet:
            print('Debet:', i)
        print(line)
        for i in kredit:
            print('Kredit:', i)
        print(line)

    def GetData():
        """Collects the data (debet, kredit) from the FilePath xlsx document's saldobalance"""

        # Try load workbook
        try:
            global wb
            wb = openpyxl.load_workbook(FileToLoad)
        except (FileNotFoundError, PermissionError) as e:
            print(e)
            return 'Could not find or load the specified document'

        # Try to get saldobalance
        try:
            saldo = wb.get_sheet_by_name(saldobalance)
        except (NameError, KeyError) as e:
            print(e)
            return "Der findes ikke noget '{}' ark.\n(Den er case sensitive - dobbelttjek store og små bogstaver)".format(
                saldobalance)

        # get data fra "saldobalance"
        global kredit
        kredit = []
        global debet
        debet = []

        for i in range(int(StartSaldoRow[1:]), int(EndSaldoRow[1:])):  # Bogstaver hvis 'letter' er A:
            if saldo[chr(dec + 2) + str(i)].value == None:  # C
                kredit.append([saldo[chr(dec) + str(i)].value,  # A
                               saldo[chr(dec + 1) + str(i)].value,  # B
                               saldo[chr(dec + 3) + str(i)].value,  # D
                               "'{}'!{}".format(saldobalance, chr(dec + 3) + str(i))])  # D
            else:
                debet.append([saldo[chr(dec) + str(i)].value,  # A
                              saldo[chr(dec + 1) + str(i)].value,  # B
                              saldo[chr(dec + 2) + str(i)].value,  # C
                              "'{}'!{}".format(saldobalance, chr(dec + 2) + str(i))])  # D

    def headlineFormula(i, headline):
        """Laver formel til headline, og gør den bold"""

        if headline == 'Bruttofortjeneste':
            doc['A5'].font = bold10
            doc['B5'] = "=B3-B4"
            return True
        elif headline == 'Resultat før primær drift':
            doc['A8'].font = bold10
            doc['B8'] = "=B5-B6-B7"
            return True
        elif headline == 'Resultat før skat':
            doc['A11'].font = bold10
            doc['B11'] = "=B8+B9-B10"
            return True
        elif headline == 'Årets resultat':
            doc['A13'].font = bold10
            doc['B13'] = "=B11-B12"
            return True

    def lavNote(count, headline, noteRow, noteNR, i, start, end, type):
        """Laver en note - increases noteRow with 2, and noteNR with 1"""

        # Overskrift
        doc['A' + str(noteRow)] = "Note {} - {}".format(noteNR, headline)
        doc['A' + str(noteRow)].font = bold10
        noteRow += 1

        for konto in type:  # type = kredit/debet list
            if start <= konto[0] < end:  # append til note hvis korrekt kontonr
                doc['A' + str(noteRow)] = konto[1]
                doc['B' + str(noteRow)] = '=' + konto[3]
                noteRow += 1

        doc['A' + str(noteRow)] = "{} i alt".format(headline)  # sum af note
        doc['B' + str(noteRow)] = "=+SUM(B{}:B{})".format(noteRow - count, noteRow - 1)
        doc['B' + str(noteRow)].font = underline10

        # Add notens sum til den tilhørende headline
        doc['B' + str(3 + i)] = '=+B{}'.format(noteRow)

        # Add noteNR ved siden af den tilhørende headline og sum
        doc['C' + str(3 + i)] = noteNR

        noteRow += 2
        noteNR += 1
        return noteRow, noteNR

    def addFormel(headline, noteRow, noteNR, i, start, end, type):
        """For at undgå gentagelse - add'er til cell, hvis der skal addes
        over 3 så lav en note"""
        formula = '='
        count = 0
        for konto in type:
            if start <= konto[0] < end:
                formula += '+' + konto[3]
                count += 1

        if count >= 3:  # lav note
            noteRow, noteNR = lavNote(count, headline, noteRow, noteNR, i, start, end, type)
            return noteRow, noteNR

        if formula != '=':  # hvis der er mindst en konto added
            doc['B' + str(3 + i)] = formula
        else:
            doc['B' + str(3 + i)] = 0

        return noteRow, noteNR

    def createFRO(title='Funktionsopd. Resultatopgørelse'):
        """Laver den funktionsopdelte resultatopgørelse"""

        # Lav ny "sheet" og gå ind på den
        wb.create_sheet(title=title, index=0)
        global doc
        doc = wb.get_sheet_by_name(title)

        # Overskrift
        doc['A1'] = 'Funktionsopdelt Resultatsopgørelse'
        doc['A1'].font = bold22Font

        headlines = (
            'Nettoomsætning',
            'Produktionsomkostninger',
            'Bruttofortjeneste',
            'Distributionsomkostninger',
            'Administrationsomkostninger',
            'Resultat før primær drift',
            'Finansielle indtægter',
            'Finansielle omkostninger',
            'Resultat før skat',
            'Skat',
            'Årets resultat')

        noteRow = 15
        noteNR = 1

        # Lav resultatopgørelse skabelon
        for i, headline in enumerate(headlines):
            doc['A' + str(3 + i)] = headline

            # Hvis headline er iblandt "overskrifterne" - bliver de behandlet nu - skip
            if headlineFormula(i, headline):
                continue

            # add formula - og lav nav note hvis 3 eller mere
            if headline == 'Nettoomsætning':
                noteRow, noteNR = addFormel(headline, noteRow, noteNR, i, start=0, end=1200, type=kredit)

            elif headline == 'Produktionsomkostninger':
                noteRow, noteNR = addFormel(headline, noteRow, noteNR, i, start=2000, end=3000, type=debet)

            elif headline == 'Distributionsomkostninger':
                noteRow, noteNR = addFormel(headline, noteRow, noteNR, i, start=3000, end=4000, type=debet)

            elif headline == 'Administrationsomkostninger':
                noteRow, noteNR = addFormel(headline, noteRow, noteNR, i, start=4000, end=5000, type=debet)

            elif headline == 'Finansielle indtægter':
                noteRow, noteNR = addFormel(headline, noteRow, noteNR, i, start=5000, end=6000, type=kredit)

            elif headline == 'Finansielle omkostninger':
                noteRow, noteNR = addFormel(headline, noteRow, noteNR, i, start=6000, end=7000, type=debet)

            elif headline == 'Skat':
                noteRow, noteNR = addFormel(headline, noteRow, noteNR, i, start=7000, end=9000, type=debet)





    status = GetData()

    if status:
        return status

    PrintKontis()
    status = createFRO()


    doc.column_dimensions['A'].width = 40

    # Save file SaveLocation
    try:
        wb.save(SaveLocation)
    except (FileNotFoundError, PermissionError) as e:
        print(e)
        return 'Invalid save location\n(If you are overwriting a document, remember to close it)'

    print('-Finished-')


